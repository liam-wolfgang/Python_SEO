# 1 import pandas to download a keyword level report from semrush
import pandas as pd
# 2
import cloudscraper
from bs4 import BeautifulSoup
# 3
from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle

keywords = pd.read_excel('./assets/dmi.xlsx')

# 1 Importing the data from Semrush
low_hanging = keywords[keywords['Position'] < 15]
low_hanging_list = low_hanging.values.tolist()

dict_urls = {}
for urls in low_hanging_list:
    if urls[6] in dict_urls:
        dict_urls[urls[6]] += [[urls[0], urls[1], urls[3]]]
    else:
        dict_urls[urls[6]] = [[urls[0], urls[1], urls[3]]]


# 2 - Scraping the URLs and finding the occurrences
scraper = cloudscraper.create_scraper()

for key, values in dict_urls.items():
    print(str(key))

    html = scraper.get(key, headers={"User-agent" : "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/51.0.2704.103 Safari/537.36"})
    soup = BeautifulSoup(html.text)

    metatitle = (soup.find('title')).get_text()
    metadescription = soup.find('meta', attrs={'name': 'description'})["content"]
    h1 = [a.get_text() for a in soup.find_all('h1')]
    h2 = [a.get_text() for a in soup.find_all('h2')]
    paragraph = [a.get_text() for a in soup.find_all('p')]

    for y in values:
        metatitle_occurance = "True"
        metadescription_occurance = "True"
        h1_occurance = "True"
        h2_occurance = "True"
        paragraph_occurance = "True"

        for z in y[0].split(" "):
            if z not in str(metatitle).lower():
                metatitle_occurance = "False"

            if z not in str(metadescription).lower():
                metadescription_occurance = "False"

            if z not in (h1).lower():
                h1_occurance = "False"

            if z not in (h2).lower():
                h2_occurance = "False"

            if z not in (paragraph).lower():
                paragraph_occurance = "False"

        y.extend([metatitle_occurance, metadescription_occurance, h1_occurance, h2_occurance, paragraph_occurance])

# 3 - Downloading as Excel file
wb=Workbook()
dest_filename = "new_ document.xlsx"
ws1 = wb.active

number=2

for key, values in dict_urls.items():
    ws1.cell(row=1, column=1).value= "URL"
    ws1.cell(row=1, column=2).value = "KEYWORD"
    ws1.cell(row=1, column=3).value = "RANKING"
    ws1.cell(row=1, column=4).value = "SEARCHES"
    ws1.cell(row=1, column=5).value = "Metatitle Occurrence"
    ws1.cell(row=1, column=6).value = "Metadescription Occurrence"
    ws1.cell(row=1, column=7).value = "H1 Occurrence"
    ws1.cell(row=1, column=8).value = "H2 Occurrence"
    ws1.cell(row=1, column=9).value = "Paragraph Occurrence"

    for list_values in values:
        ws1.cell(row=number, column=1).value=key
        column = 2
        for iteration in list_values:
            ws1.cell(row=number, column=column).value = iteration
            column += 1
        number += 1

red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
green_text = Font(color="FFFFFF")
green_fill = PatternFill(bgColor="009c48")

dxf = DifferentialStyle(font=red_text, fill=red_fill)
dxf2 = DifferentialStyle(font=green_text, fill=green_fill)

rule = Rule(type="containsText", operator="containsText", formula=['A1:N' + str(number) + '= "False"'], dxf=dxf)
rule2 = Rule(type="containsText", operator="containsText", formula=['A1:N' + str(number) + '= "True"'], dxf=dxf2)

ws1.conditional_formatting.add('A1:N' + str(number), rule)
ws1.conditional_formatting.add('A1:N' + str(number), rule2)

wb.save(filename = dest_filename)



